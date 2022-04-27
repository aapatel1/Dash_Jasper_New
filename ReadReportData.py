import re
from dateutil.parser import parse
import datetime
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from dateutil.relativedelta import relativedelta


class Load_Report:
    def read_report(self, report_name, report_data, locations, wb, output_file, report_date):
        self.report_date = report_date
        self.float_headers = ["Unearned Revenue Opening", "Invoiced", "Total", "Earned Revenue",
                              "Unearned Revenue Closing", "Billed", "Discount", "Net Billed", "Camp Price", "PMTS",
                              "EMP CR", "MAN CR", "EP DISC", "GL Total", "Amount"]
        self.int_headers = ["Classes", "Enrollment", "Games"]

        if "class" in str(report_name).lower() or "camp" in str(report_name).lower() \
                or "team" in str(report_name).lower():
            report_list = self.extract_ClassCampTeam(report_data)
            if not report_list:
                return False

            output_status = self.class_camp_team_output(output_file, report_name, report_list, locations, wb)
            return output_status

        elif "membership" in str(report_name).lower():
            report_list = self.extract_Membership(report_data)
            if not report_list:
                return False

            output_status = self.membership_output(output_file, report_name, report_list, wb)
            return output_status

        elif "event" in str(report_name).lower():
            report_list = self.extract_Event(report_data)
            if not report_list:
                return False

            output_status = self.event_output(output_file, report_name, report_list, wb, report_date)
            return output_status

        elif "season" in str(report_name).lower():
            season_list = self.extract_season(wb, report_data, output_file)
            return season_list

    def extract_ClassCampTeam(self, report_data):
        table_list = []
        reportValues = report_data.find("div", attrs={"id": "reportOutput"}).find("table").next_element. \
            find_next_siblings("tr")
        for table_row in reportValues:
            try:
                row_list = []
                row_item = table_row.find_all('td')
                for td in row_item:
                    next_ele = td.next_element.next_element
                    if next_ele.name == 'span':
                        if next_ele.text:
                            row_list.append(str(td.text).replace('\n', ''))

                if row_list:
                    if len(row_list) > 1:
                        if "Printed on" not in str(row_list[1]):
                            table_list.append(row_list)
                    else:
                        table_list.append(row_list)
            except Exception as e:
                print(str(e))

        return table_list

    def extract_Membership(self, report_data):
        table_list = []
        reportValues = report_data.find("div", attrs={"id": "reportOutput"}).find_all("tr")
        for table_row in reportValues:
            try:
                row_list = []
                row_item = table_row.find_all('td')
                if len(row_item) > 2 and str(row_item[1].text).replace("\n", "") == "Total":
                    continue
                for td_num, td in enumerate(row_item):
                    if td.text:
                        row_list.append(str(td.text).replace('\n', ''))
                if len(row_list) > 1:
                    if "Printed on" not in str(row_list[1]):
                        table_list.append(row_list)
            except Exception as e:
                print(str(e))

        return table_list

    def extract_Event(self, report_data):
        table_list = []
        for location, reportData in report_data.items():
            reportValues = reportData.find("div", attrs={"id": "reportOutput"}).find_all("tr")
            for table_row in reportValues:
                try:
                    row_list = []
                    row_list.append(location)
                    row_item = table_row.find_all('td')
                    for td in row_item:
                        if "data-tableuuid" in str(td):
                            if td.text:
                                row_list.append(str(td.text).replace('\n', ''))

                    if len(row_list) > 2:
                        if "Printed on" not in str(row_list[1]):
                            table_list.append(row_list)
                except Exception as e:
                    print(str(e))

        return table_list

    def extract_season(self, wb, report_data, output_file):
        ws = wb.create_sheet("Seasons", -1)
        for row_num, row in enumerate(report_data):
            for col_num, value in enumerate(row):
                ws.cell(row=row_num + 1, column=col_num + 1).value = value
                if row_num > 0:
                    try:
                        if col_num == 1 or col_num == 2:
                            date = parse(str(value))
                            ws.cell(row=row_num + 1, column=col_num + 1).value = date
                            ws.cell(row=row_num + 1, column=col_num + 1).number_format = f'mm-dd-yyyy'
                    except:
                        pass

        wb.save(output_file)
        return True

    def class_camp_team_output(self, output_file, report_name, table_list, locations, wb):
        location = start_date = end_date = None
        output_list = []
        if "class" in str(report_name).lower():
            headers = [
                "Location", "Start Date", "End Date", "Season Name", "GL Code", "Product Name", "Level Name", "Month",
                "Classes", "Unearned Revenue Opening", "Invoiced", "Total", "Earned Revenue", "Unearned Revenue Closing"
            ]
            output_list.append(headers)
        else:
            headers = table_list[2]
            headers.insert(0, "Location")
            headers.insert(1, "Start Date")
            headers.insert(2, "End Date")
            output_list.append(headers)

        for line_num, line in enumerate(table_list[3:]):
            seasonName = str(line[0])
            new_row = []
            LOCATION_ROW = False
            DATE_ROW = False

            if seasonName == "Total" or seasonName == '':
                continue

            for locName, locValue in locations.items():
                if locName in seasonName:
                    location = locName
                    LOCATION_ROW = True
                    break

            if "[" in seasonName:
                start_date = str(seasonName).split('[')[1].split(' - ')[0].strip()
                end_date = str(seasonName).split('[')[1].split(' - ')[1].replace(']', '').strip()
                DATE_ROW = True

            if LOCATION_ROW or DATE_ROW:
                continue

            if location and start_date and end_date:
                new_row.insert(0, location)
                new_row.insert(1, start_date)
                new_row.insert(2, end_date)

                for val in line:
                    new_row.append(val)
                output_list.append(new_row)

        report = str(report_name).replace('Unearned Revenue -', '').strip()
        ws = wb[report]
        for row_num, row in enumerate(output_list):
            for col_num, value in enumerate(row):
                if row_num == 0:
                    ws.cell(row=row_num + 1, column=col_num + 1).value = value

                elif output_list[0][col_num] in self.int_headers:
                    try:
                        value = str(value).replace("$", "").replace(",", "").replace("- ", "-").strip() or 0
                        value = float(value)
                        ws.cell(row=row_num + 1, column=col_num + 1).value = value
                    except:
                        ws.cell(row=row_num + 1, column=col_num + 1).value = value

                elif output_list[0][col_num] in self.float_headers:
                    try:
                        value = str(value).replace("$", "").replace(",", "").replace("- ", "-").strip() or 0
                        value = float(value)
                        ws.cell(row=row_num + 1, column=col_num + 1).value = value
                        ws.cell(row=row_num + 1, column=col_num + 1).number_format = f'$#,##0.00_);[Red]($#,##0.00)'
                    except:
                        ws.cell(row=row_num + 1, column=col_num + 1).value = value

                else:
                    ws.cell(row=row_num + 1, column=col_num + 1).value = value

                try:
                    date = parse(value)
                    if col_num == 1 or col_num == 2:
                        ws.cell(row=row_num + 1, column=col_num + 1).value = date
                        ws.cell(row=row_num + 1, column=col_num + 1).number_format = f'mm-dd-yyyy'

                    if col_num == 2:
                        if date <= parse(self.report_date):
                            ws.cell(row=row_num + 1, column=col_num + 1).fill = \
                                PatternFill(start_color='FFF0F000', end_color='FFF0F000', fill_type="solid")
                except:
                    pass

        wb.save(output_file)
        return True

    def membership_output(self, output_file, report_name, table_list, wb):
        report = str(report_name).replace('Unearned Revenue - ', '').strip()
        ws = wb[report]
        for row_num, row in enumerate(table_list[5:]):
            for col_num, value in enumerate(row[1:]):
                if col_num > 1:
                    try:
                        value = str(value).replace("$", "").replace(",", "").replace("- ", "-").strip()
                        value = float(value)
                        ws.cell(row=row_num + 1, column=col_num + 1).value = value
                        ws.cell(row=row_num + 1, column=col_num + 1).number_format = f'$#,##0.00_);[Red]($#,##0.00)'
                    except:
                        ws.cell(row=row_num + 1, column=col_num + 1).value = value
                else:
                    ws.cell(row=row_num + 1, column=col_num + 1).value = value

        wb.save(output_file)
        return True

    def event_output(self, output_file, report_name, table_list, wb, report_date):
        report = str(report_name).replace('Unearned Revenue - ', '').strip()
        ws = wb[report]
        headers = table_list[0]
        headers.append("GL Total")
        header_check = ['GL Code', 'Invoice Date', 'Amount']

        glCode = glDes = glTotal = 0
        eventList = []
        eventList.append(headers)
        for line in table_list:
            if set(header_check).issubset(line):
                continue
            if 'Grand Total' in line or '©2018 DASH Platform by SportsIT' in line:
                continue
            if any(len(item) > 0 for item in line):
                if line[2]:
                    glCode = line[1]
                    glDes = line[2]
                    glTotal = str(line[8]).replace("$", "").replace(",", "").replace("- ", "-").strip() or 0
                    glTotal = float(glTotal)
                else:
                    location = line[0]
                    custName = line[6]
                    eventDes = line[7]
                    eventDate = line[8]
                    invoice = line[9]
                    invoiceDate = line[10]
                    amount = str(line[11]).replace("$", "").replace(",", "").replace("- ", "-").strip() or 0
                    amount = float(amount)
                    eventList.append([location, glCode, glDes, custName, eventDes, eventDate, invoice, invoiceDate,
                                      amount, glTotal])
        eventList[0][0] = "Location"

        for row_num, row in enumerate(eventList):
            HIGHLIGHT = False
            for col_num, value in enumerate(row):
                event_date = row[5] if row_num > 0 else None

                if eventList[0][col_num] in self.float_headers:
                    try:
                        value = str(value).replace("$", "").replace(",", "").replace("- ", "-").strip()
                        value = float(value)
                        ws.cell(row=row_num + 1, column=col_num + 1).value = value
                        ws.cell(row=row_num + 1, column=col_num + 1).number_format = f'$#,##0.00_);[Red]($#,##0.00)'
                    except:
                        ws.cell(row=row_num + 1, column=col_num + 1).value = value

                else:
                    if event_date and not HIGHLIGHT:
                        event_date = parse(str(event_date))
                        report_date_ = parse(str(report_date))
                        if report_date_ >= event_date:
                            HIGHLIGHT = True

                    ws.cell(row=row_num + 1, column=col_num + 1).value = value
                    if HIGHLIGHT:
                        ws.cell(row=row_num + 1, column=col_num + 1).value = value
                        ws.cell(row=row_num + 1, column=col_num + 1).fill = \
                            PatternFill(start_color='FFFCE4D6', end_color='FFFCE4D6', fill_type="solid")

                    try:
                        if col_num == 5 or col_num == 7:
                            date = parse(value)
                            ws.cell(row=row_num + 1, column=col_num + 1).value = date
                            ws.cell(row=row_num + 1, column=col_num + 1).number_format = f'mm-dd-yyyy'
                    except:
                        pass

        wb.save(output_file)
        return True
