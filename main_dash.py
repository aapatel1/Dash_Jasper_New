from dateutil.parser import parse
import aiohttp
import asyncio
import json
import time
from bs4 import BeautifulSoup as bs
import os
import binascii
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook, load_workbook
from dash_requests import DashRequests
from ReadReportData import Load_Report

LIMIT = 5
TIMEOUT = 600  # seconds
USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 ' \
             'Safari/537.36 '


class DASH_Download:
    def __init__(self, client_name, report_date, output_path, all_reports, output_file):
        self.gui_queue = None
        self.username = self.password = self.company_id = None
        self.client_name = client_name
        self.other_month = False
        self.report_date = report_date
        self.download_path = output_path
        self.output_file = output_file
        self.all_reports = all_reports
        self.wb = None

    def gui_queue_status(self, msg):
        self.gui_queue.put({'status': f'{msg}'}) if self.gui_queue else None

    async def process_dash_report(self):
        timeout = aiohttp.ClientTimeout(total=TIMEOUT)
        conn = aiohttp.TCPConnector(limit=5, limit_per_host=5)
        self.sema = asyncio.Semaphore(LIMIT)
        async with aiohttp.ClientSession(connector=conn, timeout=timeout) as self.session:
            dash = DashRequests(self.sema, self.session)
            dash_login = await dash.dash_login(self.company_id, self.username, self.password)
            if not dash_login:
                msg = f"Login Failed : {self.client_name}"
                self.gui_queue_status(msg)
                return False

            if isinstance(dash_login, bool):
                msg = f"Login Successfully : {self.client_name}"
                self.gui_queue_status(msg)
            else:
                self.gui_queue_status(dash_login)
                return False

            lr = Load_Report()
            dash_locations = await dash.get_locations()
            dash_all_searches = await dash.dash_all_search()
            REPORT_FOUND = False
            for report_num, report in enumerate(self.all_reports, 1):
                for all_search in dash_all_searches:
                    if all_search.get('search').strip() == str(report).strip():
                        REPORT_FOUND = True
                        report_url = all_search.get('url')
                        jasper_code = await dash.jasper_login()
                        report_view = await dash.report_view(report_url)
                        load_report = await dash.load_report(report, report_url)
                        input_control = await dash.input_controls(report_url)
                        report_value = await dash.report_values(report_url, input_control)
                        if str(report) != "Unearned Revenue - Event & Rental":
                            report_flow = await dash.report_flow(report, report_num, self.report_date,
                                                                 self.other_month)
                        else:
                            report_flow = self.event_location_report(dash_locations)

                        if report_flow:
                            report_sheet_name = str(report).split('-')[1].strip()
                            self.wb.create_sheet(report_sheet_name)
                            lr.read_report(report, report_flow, dash_locations, self.wb, self.output_file,
                                           self.report_date)
                            status = f'{self.client_name} {report} : Download Successfully.'
                            print(status)
                            self.gui_queue_status(status)
                        else:
                            status = f'{self.client_name} {report} : Download Failed.'
                            print(status)
                            self.gui_queue_status(status)

                if not REPORT_FOUND:
                    msg = f'Report not Found : {report}'
                    self.gui_queue_status(msg)

        return True

    def event_location_report(self, dash_locations):
        pass

    def download_process(self):
        loop = asyncio.new_event_loop()
        # self.executor = ThreadPoolExecutor(max_workers=3)
        future = asyncio.ensure_future(self.process_dash_report(), loop=loop)
        loop.run_until_complete(future)


class run_download_report:
    def run_download(self, gui_queue, report_date, output_path, client_name, all_reports):
        print('Starting process for', client_name)
        start_time = time.perf_counter()
        self.setting_xl = 'Dash_SettingSheet.xlsx'
        self.setting_wb = load_workbook(self.setting_xl)
        creds_values = self.setting_wb['Creds']
        credentials = [list(rows) for rows in creds_values.values]

        output_path = os.path.join(output_path, "Downloads", client_name)
        if not os.path.exists(output_path):
            os.makedirs(output_path)

        reportDate = str(report_date).replace('/', '-')
        output_file = os.path.join(output_path, f'Unearned Revenue - Summary {reportDate}.xlsx')
        report_date = parse(report_date).strftime("%Y-%m-%d")

        dash = DASH_Download(client_name, report_date, output_path, all_reports, output_file)
        dash.GET_SEASON = True
        dash.gui_queue = gui_queue

        if not os.path.isfile(output_file):
            dash.wb = Workbook()
            dash.ws = dash.wb.active
            dash.ws.title = "Summary"
            dash.wb.save(output_file)
        else:
            dash.wb = load_workbook(output_file)
            dash.ws = dash.wb.active

        dash.company_id = str([str(item[2]).strip() for item in credentials if
                              str(item[0]).strip().lower() == str(client_name).strip().lower()][0])
        dash.username = str([str(item[3]).strip() for item in credentials if
                             str(item[0]).strip().lower() == str(client_name).strip().lower()][0])
        dash.password = str([str(item[4]).strip() for item in credentials if
                             str(item[0]).strip().lower() == str(client_name).strip().lower()][0])

        dash.download_process()

        end_time = time.perf_counter()
        time_taken = time.strftime("%H:%M:%S", time.gmtime(int(end_time - start_time)))
        gui_queue.put({'status': f'\n\nTime Taken : {time_taken}'}) if gui_queue else None
        return True
