from openpyxl import load_workbook, Workbook
from time import sleep
import os
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from dateutil.parser import parse
from datetime import datetime


class SplitSummary:
    def __init__(self):
        self.camp_amt_column = self.class_amt_column = self.event_amt_column = None
        self.camp_header = self.class_header = self.event_header = None
        self.report_date = self.output_path = self.client = None

    def load_reports(self, downloaded_reports):
        self.report_month = parse(self.report_date).strftime("%B %Y")
        load_reports = load_workbook(downloaded_reports, read_only=True, data_only=True)
        report_sheet_names = load_reports.sheetnames
        self.report_dict = {}
        for sheet_name in report_sheet_names:
            self.location_dict = {}
            reports_sheet = load_reports[sheet_name]
            if sheet_name in ["Camp", "Class"]:
                report_data = [row for row in reports_sheet.values]
                header_row = report_data[0]
                self.get_amount_col(sheet_name, header_row)
                self.break_locations(report_data)
                self.report_dict[sheet_name] = self.location_dict
            if sheet_name == "Event & Rental":
                report_data = [row for row in reports_sheet.values]
                header_row = report_data[0]
                self.get_amount_col(sheet_name, header_row)
                self.break_event_locations(report_data)
                self.report_dict[sheet_name] = self.location_dict

        self.prepare_output()
        return True

    def get_amount_col(self, sheet_name, header_row):
        for col_num, value in enumerate(header_row):
            if sheet_name == "Camp":
                self.camp_header = header_row
                if value == "Unearned Revenue Closing":
                    self.camp_amt_column = col_num
            elif sheet_name == "Class":
                self.class_header = header_row
                if value == "Unearned Revenue Closing":
                    self.class_amt_column = col_num
            elif sheet_name == "Event & Rental":
                self.event_header = header_row
                if value == "Amount":
                    self.event_amt_colnum = col_num

    def gl_season_break_locations(self, report_data):
        self.locations = list(set([item[0] for item in report_data[1:]]))
        for location in self.locations:
            gl_list = list(set([item[4] for item in report_data[1:] if location in item]))
            self.season_dict = {}
            for gl in gl_list:
                season_report_data = [row for row in report_data if gl in row and location in row]
                seasons = [item[3] for item in season_report_data if item]
                for season in seasons:
                    season_data = [row for row in season_report_data if season in row]
                    self.season_dict[gl, season] = season_data
            self.location_dict[location] = self.season_dict
        return True

    def break_locations(self, report_data):
        self.locations = list(set([item[0] for item in report_data[1:]]))
        for location in self.locations:
            self.season_dict = {}
            season_report_data = [row for row in report_data if location in row]
            seasons = [item[3] for item in season_report_data if item]
            for season in seasons:
                season_data = [row for row in season_report_data if season in row]
                self.season_dict[season] = season_data
            self.location_dict[location] = self.season_dict
        return True

    def break_event_locations(self, report_data):
        locations = list(set([item[0] for item in report_data[1:]]))
        for location in locations:
            season_data = [item for item in report_data[1:] if location in item]
            self.location_dict[location] = season_data
        return True

    def prepare_output(self):
        for location in self.locations:
            summary_dict = {}
            output_file_name = f'{location} - {self.report_month} Unearned Revenue Working.xlsx'
            output_file = os.path.join(self.output_path, output_file_name)
            wb = Workbook()
            summary_ws = wb.active
            summary_ws.title = "Summary"

            for report_name, location_dict in self.report_dict.items():
                if report_name in ["Camp", "Class"]:
                    amt_col = self.camp_amt_column if report_name == "Camp" else self.class_amt_column
                    for location_name, season_dict in location_dict.items():
                        if location_name == location:
                            for num, (season_name, season_data) in enumerate(season_dict.items(), 1):
                                for row in season_data:
                                    amount = row[amt_col]
                                    if not amount:
                                        continue

                                    sheet_code = "CA" if report_name == "Camp" else "CL"
                                    sheet_code = f"{sheet_code}{str(num)}"
                                    if sheet_code not in wb.sheetnames:
                                        new_sheet = wb.create_sheet(sheet_code, num)
                                        self.insert_headers(report_name, new_sheet)
                                    else:
                                        new_sheet = wb[sheet_code]

                                    new_sheet.append(row[3:])
                                    gl_code = row[4]
                                    if not summary_dict or (gl_code, season_name, sheet_code) not in summary_dict:
                                        summary_dict[gl_code, season_name, sheet_code] = round(float(amount), 2)
                                    else:
                                        summary_dict[gl_code, season_name, sheet_code] += round(float(amount), 2)

                if report_name == "Event & Rental":
                    amt_col = self.event_amt_colnum
                    for location_name, season_list in location_dict.items():
                        if location_name == location:
                            new_sheet = wb.create_sheet("E1")
                            self.insert_headers(report_name, new_sheet)
                            for row in season_list:
                                new_sheet.append(row[1:])
                                gl_code = row[1]
                                gl_description = row[2]
                                amount = round(float(row[amt_col]), 2)
                                if not summary_dict or (gl_code, gl_description, "E1") not in summary_dict:
                                    summary_dict[gl_code, gl_description, "E1"] = round(float(amount), 2)
                                else:
                                    summary_dict[gl_code, gl_description, "E1"] += round(float(amount), 2)

            summary_list = self.prepare_summary_list(summary_dict)
            self.append_summary(summary_ws, summary_list)
            wb.save(output_file)

    def prepare_summary_list(self, summary_dict):
        camp_dict = {}
        class_dict = {}
        event_dict = {}
        for keys, value in summary_dict.items():
            if keys[2].startswith("CA"):
                if not camp_dict:
                    camp_dict["Camp - Unearned Revenue"] = []
                camp_dict["Camp - Unearned Revenue"].append([float(keys[0]), keys[1], value, keys[2]])
            if keys[2].startswith("CL"):
                if not class_dict:
                    class_dict["Class - Unearned Revenue"] = []
                class_dict["Class - Unearned Revenue"].append([float(keys[0]), keys[1], value, keys[2]])
            if keys[2].startswith("E"):
                if not event_dict:
                    event_dict["Event - Unearned Revenue"] = []
                event_dict["Event - Unearned Revenue"].append([float(keys[0]), keys[1], value, keys[2]])

        summary_list = []
        if class_dict:
            summary_list.append(class_dict)
        if camp_dict:
            summary_list.append(camp_dict)
        if event_dict:
            summary_list.append(event_dict)

        return summary_list

    def append_summary(self, summary_ws, summary_list):
        summary_ws.cell(row=2, column=2).value = f"Unearned Revenue Summary as of {self.report_date}"
        summary_ws.merge_cells(start_row=2, end_row=2, start_column=2, end_column=5)
        self.cell_formatting(summary_ws, row_num=2, col_num=2,
                             font_name="Calibri", font_bold=True, font_underline=None, font_size=18, font_color="FFFFFF",
                             alignment_horizontal="center", alignment_vertical="center",
                             pfill_start_color="FF00B0F0", pfill_end_color="FF00B0F0", pfill_type="solid")

        summary_total = 0
        for summary_dict in summary_list:
            for reportName, reportData in summary_dict.items():
                total_row = summary_ws.max_row + 3
                summary_ws.cell(row=total_row, column=3).value = f"{reportName}"
                self.cell_formatting(summary_ws, row_num=total_row, col_num=3,
                                     font_name="Calibri", font_bold=True, font_underline='single', font_size=14, font_color=None,
                                     alignment_horizontal="center", alignment_vertical="center",
                                     pfill_start_color=None, pfill_end_color=None, pfill_type=None)

                total_row = summary_ws.max_row + 1
                summary_ws.cell(row=total_row, column=2).value = f"GL Code"
                self.cell_formatting(summary_ws, row_num=total_row, col_num=2,
                                     font_name="Calibri", font_bold=True, font_underline=None, font_size=11, font_color=None,
                                     alignment_horizontal="center", alignment_vertical="center",
                                     pfill_start_color="FFFCE4D6", pfill_end_color="FFFCE4D6", pfill_type="solid")
                self.cell_borders_4_single(summary_ws, total_row, 2)

                summary_ws.cell(row=total_row, column=3).value = f"Name"
                self.cell_formatting(summary_ws, row_num=total_row, col_num=3,
                                     font_name="Calibri", font_bold=True, font_underline=None, font_size=11, font_color=None,
                                     alignment_horizontal="center", alignment_vertical="center",
                                     pfill_start_color="FFFCE4D6", pfill_end_color="FFFCE4D6", pfill_type="solid")
                self.cell_borders_4_single(summary_ws, total_row, 3)

                summary_ws.cell(row=total_row, column=4).value = f"Amount"
                self.cell_formatting(summary_ws, row_num=total_row, col_num=4,
                                     font_name="Calibri", font_bold=True, font_underline=None, font_size=11, font_color=None,
                                     alignment_horizontal="center", alignment_vertical="center",
                                     pfill_start_color="FFFCE4D6", pfill_end_color="FFFCE4D6", pfill_type="solid")
                self.cell_borders_4_single(summary_ws, total_row, 4)

                summary_ws.cell(row=total_row, column=5).value = f"No."
                self.cell_formatting(summary_ws, row_num=total_row, col_num=5,
                                     font_name="Calibri", font_bold=True, font_underline=None, font_size=11, font_color=None,
                                     alignment_horizontal="center", alignment_vertical="center",
                                     pfill_start_color="FFFCE4D6", pfill_end_color="FFFCE4D6", pfill_type="solid")
                self.cell_borders_4_single(summary_ws, total_row, 5)

                total_row = summary_ws.max_row
                total_amount = 0
                for row_num, row in enumerate(reportData, total_row):
                    for col_num, value in enumerate(row, 2):
                        summary_ws.cell(row=row_num+1, column=col_num).value = value
                        summary_ws.cell(row=row_num+1, column=col_num).alignment = \
                            Alignment(horizontal='center', vertical='center')
                        self.cell_borders_4_single(summary_ws, row_num+1, col_num)
                        if col_num == 4:
                            summary_ws.cell(row=row_num+1, column=col_num).number_format = \
                                f'$#,##0.00_);[Red]($#,##0.00)'
                            total_amount += float(value)
                        if col_num == 5:
                            summary_ws.cell(row=row_num+1, column=col_num).hyperlink = f'#{value}!A1'
                            self.cell_formatting(summary_ws, row_num=row_num+1, col_num=col_num,
                                                 font_name="Calibri", font_bold=False, font_underline='single',
                                                 font_size=11, font_color="FF5672C4",
                                                 alignment_horizontal="center", alignment_vertical="center",
                                                 pfill_start_color=None, pfill_end_color=None, pfill_type=None)

                summary_total += total_amount
                summary_ws.append([])
                total_row = summary_ws.max_row
                for col in range(2, 6):
                    self.cell_borders_4_single(summary_ws, total_row+1, col)

                summary_ws.cell(row=total_row+2, column=3).value = "Total"
                summary_ws.cell(row=total_row+2, column=3).font = Font(name="Calibri", bold=True, size=11)
                summary_ws.cell(row=total_row+2, column=3).alignment = Alignment(horizontal='center', vertical='center')
                for col in range(2, 6):
                    self.cell_borders_4_single(summary_ws, total_row+2, col)
                self.cell_borders_total_cell(summary_ws, total_row+2, 3)

                summary_ws.cell(row=total_row+2, column=4).value = total_amount
                summary_ws.cell(row=total_row+2, column=4).number_format = f'$#,##0.00_);[Red]($#,##0.00)'
                summary_ws.cell(row=total_row+2, column=4).font = Font(name="Calibri", bold=True, size=11)
                self.cell_borders_total_cell(summary_ws, total_row+2, 4)
                summary_ws.append([])
                summary_ws.append([])

        total_row = summary_ws.max_row
        summary_ws.cell(row=total_row+2, column=3).value = "TOTAL :"
        summary_ws.cell(row=total_row+2, column=3).font = Font(name="Calibri", bold=True, size=11)
        summary_ws.cell(row=total_row+2, column=3).alignment = Alignment(horizontal='center', vertical='center')

        self.cell_borders_total_cell(summary_ws, total_row+2, 3)
        summary_ws.cell(row=total_row+2, column=4).value = summary_total
        summary_ws.cell(row=total_row+2, column=4).number_format = f'$#,##0.00_);[Red]($#,##0.00)'
        summary_ws.cell(row=total_row+2, column=4).font = Font(name="Calibri", bold=True, size=11)
        self.cell_borders_total_cell(summary_ws, total_row+2, 4)

        summary_ws.column_dimensions[get_column_letter(2)].width = 10
        summary_ws.column_dimensions[get_column_letter(3)].width = 50
        summary_ws.column_dimensions[get_column_letter(4)].width = 15
        summary_ws.column_dimensions[get_column_letter(5)].width = 10
        summary_ws.sheet_view.showGridLines = False
        return True

    def cell_formatting(self, summary_ws, row_num, col_num, font_name, font_bold, font_underline, font_size, font_color,
                        alignment_horizontal, alignment_vertical, pfill_start_color, pfill_end_color, pfill_type):
        summary_ws.cell(row=row_num, column=col_num).font = \
            Font(name=font_name, bold=font_bold, underline=font_underline, size=font_size, color=font_color)
        summary_ws.cell(row=row_num, column=col_num).alignment = \
            Alignment(horizontal=alignment_horizontal, vertical=alignment_vertical)
        summary_ws.cell(row=row_num, column=col_num).fill = \
            PatternFill(start_color=pfill_start_color, end_color=pfill_end_color, fill_type=pfill_type)

    def cell_borders_4_single(self, summary_ws, row_num, col_num):
        summary_ws.cell(row=row_num, column=col_num).border = \
            Border(top=Side(border_style='thin', color='FF000000'),
                   right=Side(border_style='thin', color='FF000000'),
                   left=Side(border_style='thin', color='FF000000'),
                   bottom=Side(border_style='thin', color='FF000000'))

    def cell_borders_total_cell(self, summary_ws, row_num, col_num):
        summary_ws.cell(row=row_num, column=col_num).border = \
            Border(top=Side(border_style='thin', color='FF000000'),
                   right=Side(border_style='thin', color='FF000000'),
                   left=Side(border_style='thin', color='FF000000'),
                   bottom=Side(border_style='double', color='FF000000'))

    def insert_headers(self, report_name, new_sheet):
        if report_name == "Camp":
            new_sheet.append(self.camp_header[3:])
        elif report_name == "Class":
            new_sheet.append(self.class_header[3:])
        elif report_name == "Event & Rental":
            new_sheet.append(self.event_header[1:])
        return True


class run_split_summary:
    def __init__(self):
        self.gui_queue = None

    def run(self, report_date, output_path, client, summary_file):
        try:
            ss = SplitSummary()
            ss.report_date = report_date
            ss.output_path = output_path
            ss.client = client
            ss.load_reports(summary_file)
            self.gui_queue.put({'status': f'\nSplit Summary Processed.'}) if self.gui_queue else None
            return True
        except:
            self.gui_queue.put({'status': f'\nFailed.'}) if self.gui_queue else None
            return False
