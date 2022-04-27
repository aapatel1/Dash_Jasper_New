from dateutil.parser import parse
import aiohttp
import asyncio
import json
import time
from bs4 import BeautifulSoup as bs
import os
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook, load_workbook
from ReadReportData import Load_Report
from datetime import datetime

LIMIT = 5
TIMEOUT = 600  # seconds
base_url = r'https://apps.daysmartrecreation.com/dash/admin/index.php'
USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 ' \
             'Safari/537.36 '

action_dict = {
    "camp": "Jasper/report/public-Reports-Financial-Unearned_Revenue__Camp_Staging_v_1",
    "class": "Jasper/report/public-Reports-Financial-Unearned_Revenue___Class_v_2",
    "event": "Jasper/report/public-Reports-Financial-Unearned_Revenue___Event",
    "membership": "Jasper/report/public-Reports-Financial-Unearned_Revenue___Membership",
    "team": "Jasper/report/public-Reports-Financial-Unearned_Revenue_Team_v_1"
}


class DASH_Download():
    def __init__(self):
        self.username = self.password = self.companyid = None
        self.sema = self.session = None
        self.client_name = self.report_name = None
        self.report_date = self.download_path = None
        self.all_searches = self.all_reports = None
        self.CSRFTOKEN = None
        self.LOGIN_FAILED = False
        self.wb = self.ws = None
        self.GET_SEASON = None
        self.report_values_list = None
        self.inputControl_params = None
        self.referer_url = None
        self.locations = self.seasons = self.eventType = self.glCodes = None
        self.output_file = None
        self.seasonsList = []
        self.season_header = False
        self.error_log = None

    # login to website
    async def login(self):
        login_url = f'{base_url}'
        headers = {
            'Connection': 'keep-alive',
            'Cache-Control': 'max-age=0',
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="99", "Microsoft Edge";v="99"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'DNT': '1',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': USER_AGENT,
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;'
                      'q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-User': '?1',
            'Sec-Fetch-Dest': 'document',
            'Accept-Language': 'en-US,en;q=0.9',
        }
        async with self.sema:
            async with self.session.get(login_url, headers=headers) as request:
                response = await request.content.read()
                content = response.decode('utf-8')
                html_content = bs(content, "html.parser")
                login_data_name = html_content.find('button', attrs={'id': 'loginButton'}).get('name')
                login_data_value = html_content.find('button', attrs={'id': 'loginButton'}).get('value')

        data = {
            '_method': 'POST',
            'company_code': self.companyid,
            'username': self.username,
            'password': self.password,
            login_data_name: login_data_value
        }

        loop_count = 0
        async with self.sema:
            while loop_count < 3:
                await asyncio.sleep(3)
                try:
                    async with self.session.post(login_url, headers=headers, data=str(data)) as request:
                        response = await request.content.read()
                        self.cookies = request.cookies
                        content = response.decode('utf-8')
                        login_page_html = bs(content, 'html.parser')
                        if 'Invalid username or password' in str(login_page_html):
                            return False
                        elif 'DaySmart Rec Admin' in str(login_page_html.find('title')):
                            return True
                        else:
                            return False
                except:
                    loop_count += 1
            return False

    async def all_job_search(self):
        url = f'{base_url}'
        headers = {
            'Connection': 'keep-alive',
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="99", "Microsoft Edge";v="99"',
            'Accept': 'application/json, text/javascript, */*; q=0.01',
            'DNT': '1',
            'X-Requested-With': 'XMLHttpRequest',
            'sec-ch-ua-mobile': '?0',
            'User-Agent': USER_AGENT,
            'sec-ch-ua-platform': '"Windows"',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Dest': 'empty',
            'Accept-Language': 'en-US,en;q=0.9',
            'If-Modified-Since': 'Wed, 16 Mar 2022 06:21:34 GMT',
        }
        params = (
            ('Action', 'all/search.json'),
        )
        loop_count = 0
        async with self.sema:
            while loop_count < 3:
                await asyncio.sleep(2)
                try:
                    async with self.session.get(url, headers=headers, params=params) as request:
                        response = await request.content.read()
                        content = response.decode('utf-8')
                        self.all_searches = json.loads(str(content).lstrip('\n')).get('data')
                        if self.all_searches:
                            return True
                        return False
                except:
                    loop_count += 1
            return False

    async def all_report(self):
        url = f'{base_url}'
        headers = {
            'Connection': 'keep-alive',
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="99", "Microsoft Edge";v="99"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'Upgrade-Insecure-Requests': '1',
            'DNT': '1',
            'User-Agent': USER_AGENT,
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;'
                      'q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-User': '?1',
            'Sec-Fetch-Dest': 'document',
            'Accept-Language': 'en-US,en;q=0.9',
        }
        params = (
            ('Action', 'all/reports'),
        )
        loop_count = 0
        async with self.sema:
            while loop_count < 3:
                try:
                    async with self.session.get(url, headers=headers, params=params) as request:
                        response = await request.content.read()
                        content = response.decode('utf-8')
                        html_content = bs(content, "html.parser")
                        title = html_content.find('title').text
                        if title == "Reports - DaySmart Rec Admin":
                            return True
                        return False
                except:
                    loop_count += 1
            return False

    async def unearned_revenue(self, report_url):
        reporturl = str(report_url).replace("?Action=", "")
        headers = {
            'Connection': 'keep-alive',
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="99", "Microsoft Edge";v="99"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'DNT': '1',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': USER_AGENT,
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;'
                      'q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-User': '?1',
            'Sec-Fetch-Dest': 'document',
            'Accept-Language': 'en-US,en;q=0.9',
        }
        params = (
            ('Action', f'{reporturl}'),
        )
        loop_count = 0
        async with self.sema:
            while loop_count < 3:
                try:
                    async with self.session.get(base_url, headers=headers, params=params) as request:
                        response = await request.content.read()
                        content = response.decode('utf-8')
                        html_content = bs(content, "html.parser")
                        title = html_content.find("title").text
                        if title == "Jasper Report - DaySmart Rec Admin":
                            return True
                        return False
                except:
                    loop_count += 1
            return False

    async def report_action(self, report_url):
        self.reportAction = str(report_url).replace("?Action=", "").replace("/view", "")
        self.reportUnit = str(self.reportAction).replace("/view", "").replace("Jasper/report", "").replace("-", "/")
        headers = {
            'Connection': 'keep-alive',
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="99", "Microsoft Edge";v="99"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'Upgrade-Insecure-Requests': '1',
            'DNT': '1',
            'User-Agent': USER_AGENT,
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;'
                      'q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Dest': 'iframe',
            'Accept-Language': 'en-US,en;q=0.9',
        }
        params = (
            ('Action', self.reportAction),
            ('decorate', 'no'),
            ('theme', 'default'),
            ('reportUnit', self.reportUnit),
        )

        try_loop = 0
        while try_loop < 2:
            try:
                async with self.sema:
                    await asyncio.sleep(2)
                    async with self.session.get(base_url, headers=headers, params=params) as request:
                        self.referer_url = str(request._url)
                        response = await request.content.read()
                        content = response.decode('utf-8')
                        html_content = bs(content, "html.parser")
                        title = html_content.find("title").text
                        self.inputControl_params = dict(request._url.query)
                        return True
            except:
                try_loop += 1
        return False

    async def get_token(self):
        url = 'https://jasper.daysmartrecreation.com/jasperserver-pro/JavaScriptServlet'
        headers = {
            'Connection': 'keep-alive',
            'Content-Length': '0',
            'sec-ch-ua': '"Google Chrome";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
            'Accept': '*/*',
            'FETCH-CSRF-TOKEN': '1',
            'X-Requested-With': 'XMLHttpRequest',
            'sec-ch-ua-mobile': '?0',
            'User-Agent': USER_AGENT,
            'sec-ch-ua-platform': '"Windows"',
            'Origin': 'https://jasper.daysmartrecreation.com',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Dest': 'empty',
            'Accept-Language': 'en-US,en;q=0.9',
        }

        async with self.sema:
            await asyncio.sleep(2)
            async with self.session.post(url, headers=headers) as request:
                response = await request.content.read()
                content = response.decode('utf-8')
                self.CSRFTOKEN = str(content).split(':')[1].strip()
                return True

    async def inputControls(self):
        url = f'https://jasper.daysmartrecreation.com/jasperserver-pro/rest_v2/reports{self.reportUnit}/inputControls/'
        headers = {
            'Connection': 'keep-alive',
            'sec-ch-ua': '"Google Chrome";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
            'OWASP_CSRFTOKEN': self.CSRFTOKEN,
            'Accept-Language': 'en-US',
            'sec-ch-ua-mobile': '?0',
            'User-Agent': USER_AGENT,
            'Content-Type': 'application/json',
            'Accept': 'application/json, text/javascript, */*; q=0.01',
            'X-Requested-With': 'XMLHttpRequest, OWASP CSRFGuard Project',
            'X-Suppress-Basic': 'true',
            'sec-ch-ua-platform': '"Windows"',
            'Origin': 'https://jasper.daysmartrecreation.com',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Dest': 'empty',
        }

        data_ = '{"_flowId":["viewReportFlow"],"standAlone":["true"],"reportUnit":["-reportUnit-"],"decorate":["no"],"theme":["default"],"tk":["-tk-"]}'
        data = data_.replace("-reportUnit-", self.inputControl_params.get("reportUnit")). \
            replace("-tk-", self.inputControl_params.get("tk"))

        async with self.sema:
            await asyncio.sleep(2)
            async with self.session.post(url, headers=headers, data=data) as request:
                response = await request.content.read()
                content = response.decode('utf-8')
                json_content = json.loads(content)
                self.inputControlValues = json_content.get('inputControl')
                return True

    def get_report_values(self):
        self.locations = {}
        self.all_locations = None
        for report_items in self.inputControlValues:
            if report_items.get('id') == 'Facility' or report_items.get('id') == 'LocationID':
                for locdict in report_items.get('state').get('options'):
                    self.locations[locdict.get('label')] = locdict.get('value')
                    if str(locdict.get('label')).strip() == 'All Locations':
                        self.all_locations = str(locdict.get('value'))

        self.seasons = {}
        for report_items in self.inputControlValues:
            if report_items.get('id') == 'Season':
                for locdict in report_items.get('state').get('options'):
                    self.seasons[locdict.get('label')] = locdict.get('value')

        self.eventType = []
        for report_items in self.inputControlValues:
            if report_items.get('id') == 'EventType':
                for locdict in report_items.get('state').get('options'):
                    self.eventType.append(locdict.get('label'))

        self.glCodes = []
        for report_items in self.inputControlValues:
            if report_items.get('id') == 'GLCode':
                for locdict in report_items.get('state').get('options'):
                    self.glCodes.append(locdict.get('label'))

    async def class_camp_report(self, other_month):
        reportURL = 'https://jasper.daysmartrecreation.com/jasperserver-pro/flow.html'
        headers = {
            'Connection': 'keep-alive',
            'sec-ch-ua': '"Google Chrome";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
            'OWASP_CSRFTOKEN': self.CSRFTOKEN,
            'sec-ch-ua-mobile': '?0',
            'User-Agent': USER_AGENT,
            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
            'Accept': 'text/html, */*; q=0.01',
            'X-Requested-With': 'XMLHttpRequest, OWASP CSRFGuard Project',
            'sec-ch-ua-platform': '"Windows"',
            'Origin': 'https://jasper.daysmartrecreation.com',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Dest': 'empty',
            'Referer': str(self.referer_url),
            'Accept-Language': 'en-US,en;q=0.9',
        }
        params = (
            ('_flowExecutionKey', 'e1s1'),
            ('_flowId', 'viewReportFlow'),
            ('_eventId', 'refreshReport'),
            ('pageIndex', '0'),
            ('decorate', 'no'),
            ('confirm', 'true'),
            ('decorator', 'empty'),
            ('ajax', 'true'),
        )
        data = {
            'ParaMonth': self.report_date,
            'startDate': self.report_date,
            'Reportview': str(other_month).lower(),
            'LocationID': '~NOTHING~',
            'Season': '~NOTHING~',
            'ShowSeasonDetails': str(other_month).lower()
        }
        async with self.sema:
            await asyncio.sleep(2)
            async with self.session.post(reportURL, headers=headers, data=data, params=params) as request:
                response = await request.content.read()
                content = response.decode('utf-8')
                login_page_html = bs(content, 'html.parser')
                title = login_page_html.find("title")
                if title:
                    if "error" in str(title.text).lower():
                        return False
                return login_page_html

    async def event_report(self, locKey, locValue):
        reportURL = 'https://jasper.daysmartrecreation.com/jasperserver-pro/flow.html'
        headers = {
            'Connection': 'keep-alive',
            'sec-ch-ua': '"Google Chrome";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
            'OWASP_CSRFTOKEN': self.CSRFTOKEN,
            'sec-ch-ua-mobile': '?0',
            'User-Agent': USER_AGENT,
            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
            'Accept': 'text/html, */*; q=0.01',
            'X-Requested-With': 'XMLHttpRequest, OWASP CSRFGuard Project',
            'sec-ch-ua-platform': '"Windows"',
            'Origin': 'https://jasper.daysmartrecreation.com',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Dest': 'empty',
            'Referer': str(self.referer_url),
            'Accept-Language': 'en-US,en;q=0.9',
        }
        params = (
            ('_flowExecutionKey', 'e1s1'),
            ('_flowId', 'viewReportFlow'),
            ('_eventId', 'refreshReport'),
            ('pageIndex', '0'),
            ('decorate', 'no'),
            ('confirm', 'true'),
            ('decorator', 'empty'),
            ('ajax', 'true'),
        )
        data = {
            'AsofDate': self.report_date,
            'LocationID': str(locValue),
            'EventType': '~NOTHING~',
            'GLCode': '~NOTHING~',
            'ShowEventDetail': 'true',
            'ShowEventWithoutInvoice': 'false'
        }
        async with self.sema:
            await asyncio.sleep(2)
            async with self.session.post(reportURL, headers=headers, data=data, params=params) as request:
                try:
                    response = await request.content.read()
                    content = response.decode('utf-8')
                    login_page_html = bs(content, 'html.parser')
                    title = login_page_html.find("title")
                    if title:
                        if "error" in str(title.text).lower():
                            return locKey, False
                    try:
                        check_data = login_page_html.find("td", attrs={"class": "jrcolHeader"}).text
                        return locKey, login_page_html
                    except:
                        return locKey, "No Data"
                except:
                    return locKey, False

    async def membership_report(self):
        reportURL = 'https://jasper.daysmartrecreation.com/jasperserver-pro/flow.html'
        headers = {
            'Connection': 'keep-alive',
            'sec-ch-ua': '"Google Chrome";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
            'OWASP_CSRFTOKEN': self.CSRFTOKEN,
            'sec-ch-ua-mobile': '?0',
            'User-Agent': USER_AGENT,
            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
            'Accept': 'text/html, */*; q=0.01',
            'X-Requested-With': 'XMLHttpRequest, OWASP CSRFGuard Project',
            'sec-ch-ua-platform': '"Windows"',
            'Origin': 'https://jasper.daysmartrecreation.com',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Dest': 'empty',
            'Referer': str(self.referer_url),
            'Accept-Language': 'en-US,en;q=0.9',
        }
        params = (
            ('_flowExecutionKey', 'e1s1'),
            ('_flowId', 'viewReportFlow'),
            ('_eventId', 'refreshReport'),
            ('pageIndex', '0'),
            ('decorate', 'no'),
            ('confirm', 'true'),
            ('decorator', 'empty'),
            ('ajax', 'true'),
        )
        data = {
            'ParaAsof': self.report_date,
            'Facility': self.all_locations,
            'Month': '~NOTHING~',
            'Year': '~NOTHING~'
        }
        async with self.sema:
            await asyncio.sleep(2)
            async with self.session.post(reportURL, headers=headers, data=data, params=params) as request:
                response = await request.content.read()
                content = response.decode('utf-8')
                login_page_html = bs(content, 'html.parser')
                title = login_page_html.find("title")
                if title:
                    if "error" in str(title.text).lower():
                        return False
                return login_page_html

    async def team_report(self):
        reportURL = 'https://jasper.daysmartrecreation.com/jasperserver-pro/flow.html'

        headers = {
            'Connection': 'keep-alive',
            'sec-ch-ua': '"Google Chrome";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
            'OWASP_CSRFTOKEN': self.CSRFTOKEN,
            'sec-ch-ua-mobile': '?0',
            'User-Agent': USER_AGENT,
            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
            'Accept': 'text/html, */*; q=0.01',
            'X-Requested-With': 'XMLHttpRequest, OWASP CSRFGuard Project',
            'sec-ch-ua-platform': '"Windows"',
            'Origin': 'https://jasper.daysmartrecreation.com',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Dest': 'empty',
            'Referer': self.referer_url,
            'Accept-Language': 'en-US,en;q=0.9',
        }

        params = (
            ('_flowExecutionKey', 'e1s1'),
            ('_flowId', 'viewReportFlow'),
            ('_eventId', 'refreshReport'),
            ('pageIndex', '0'),
            ('decorate', 'no'),
            ('confirm', 'true'),
            ('decorator', 'empty'),
            ('ajax', 'true'),
        )

        data = {
            'startDate': str(self.report_date),
            'ShowSeasonDetails': 'false',
            'LocationID': '~NOTHING~',
            'Season': '~NOTHING~'
        }

        async with self.sema:
            await asyncio.sleep(2)
            async with self.session.post(reportURL, headers=headers, data=data, params=params) as request:
                response = await request.content.read()
                content = response.decode('utf-8')
                login_page_html = bs(content, 'html.parser')
                title = login_page_html.find("title")
                if title:
                    if "error" in str(title.text).lower():
                        return False
                return login_page_html

    async def get_season_list(self, year):
        headers = {
            'Connection': 'keep-alive',
            'sec-ch-ua': '"Microsoft Edge";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
            'Accept': '*/*',
            'DNT': '1',
            'X-Requested-With': 'XMLHttpRequest',
            'sec-ch-ua-mobile': '?0',
            'User-Agent': USER_AGENT,
            'sec-ch-ua-platform': '"Windows"',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Dest': 'empty',
            'Referer': 'https://apps2.daysmartrecreation.com/dash/admin/?Action=Season',
            'Accept-Language': 'en-US,en;q=0.9',
        }

        params = (
            ('Action', 'Season/show'),
            ('seasonYear', str(year)),
            ('seasonLocation', '0'),
        )

        async with self.sema:
            await asyncio.sleep(2)
            async with self.session.get(base_url, headers=headers, params=params) as request:
                response = await request.content.read()
                content = response.decode('utf-8')
                season_html = bs(content, 'html.parser')
                season_list = self.all_season_list(season_html)
                return season_list

    def all_season_list(self, season_html):
        table = season_html.find_all("tr")
        for tr in table:
            if not self.season_header:
                th_row = []
                th = tr.find_all("th")
                if th:
                    for th_value in th:
                        th_row.append(th_value.text)
                    self.seasonsList.append(th_row)
                    self.season_header = True
            td = tr.find_all("td")
            if td:
                td_row = []
                for td_value in td:
                    value = str(td_value.text).replace("\n", "").replace("\t", "").strip()
                    td_row.append(value)
                self.seasonsList.append(td_row)

        return self.seasonsList

    async def process_dash_report(self, gui_queue, report, other_month, executor):
        timeout = aiohttp.ClientTimeout(total=TIMEOUT)
        conn = aiohttp.TCPConnector(limit=5, limit_per_host=5)
        self.sema = asyncio.Semaphore(LIMIT)
        async with aiohttp.ClientSession(connector=conn, timeout=timeout) as self.session:
            dash_login = await self.login()
            if dash_login:
                print("Login Successfully")
                allSearch = await self.all_job_search()
                if not allSearch:
                    err_msg = 'Error : Report URL not found.'
                    gui_queue.put({'status': f'Error : Report URL not found.'}) if gui_queue else None
                    self.error_log.append([report, None, err_msg])
                    return False
                else:
                    for report_dict in self.all_searches:
                        if report_dict.get('search') == report:
                            report_url = report_dict.get('url')
                            print(report_url)
                            try:
                                await self.unearned_revenue(report_url)
                                await self.report_action(report_url)
                                await self.get_token()
                                if not self.CSRFTOKEN:
                                    print(f"Error : {report}")
                                    self.error_log.append([report, None, "Error"])
                                    continue

                                await self.inputControls()
                                await asyncio.sleep(5)

                                loop = asyncio.get_event_loop()
                                await loop.run_in_executor(executor, self.get_report_values)

                                if "Class" in report or "Camp" in report:
                                    report_data = await self.class_camp_report(other_month)
                                    if "The report is empty" in str(report_data):
                                        self.error_log.append([report, None, "Report is Empty"])
                                        gui_queue.put(
                                            {"status": f"\tThe report is empty: "
                                                       f"{str(report).replace('Unearned Revenue -', '').strip()}"}) \
                                            if gui_queue else None
                                        continue

                                elif "Event" in report:
                                    event_dict = {}
                                    tasks = []
                                    for locKey, locValue in self.locations.items():
                                        tasks.append(
                                            self.event_report(
                                                locKey=locKey,
                                                locValue=locValue
                                            )
                                        )
                                    if tasks:
                                        loop = asyncio.get_event_loop()
                                        for future in asyncio.as_completed(tasks, loop=loop):
                                            responses = await future
                                            location_name = responses[0]
                                            report_data = responses[1]

                                            if not report_data:
                                                self.error_log.append([report, location_name, "Error"])
                                                gui_queue.put({
                                                    "status": f"\tError : "
                                                              f"{str(report).replace('Unearned Revenue -', '').strip()}"
                                                              f" - {location_name}"}) \
                                                    if gui_queue else None
                                                continue

                                            elif "The report is empty" in str(report_data) or report_data == "No Data":
                                                self.error_log.append([report, location_name, "Report is Empty"])
                                                gui_queue.put({
                                                    "status": f"\tNo Data : "
                                                              f"{str(report).replace('Unearned Revenue -', '').strip()}"
                                                              f" - {location_name}"}) \
                                                    if gui_queue else None
                                                continue

                                            else:
                                                self.error_log.append([report, location_name, "Success"])
                                                if report_data:
                                                    event_dict[location_name] = report_data

                                elif "Membership" in report:
                                    report_data = await self.membership_report()
                                    if "The report is empty" in str(report_data):
                                        self.error_log.append([report, None, "Report is Empty"])
                                        gui_queue.put({"status":
                                                       f"\tThe report is empty: "
                                                       f"{str(report).replace('Unearned Revenue -', '').strip()}"}) \
                                            if gui_queue else None
                                        continue

                                elif "Team" in report:
                                    report_data = await self.team_report()
                                    if "The report is empty" in str(report_data):
                                        self.error_log.append([report, None, "Report is Empty"])
                                        gui_queue.put({"status":
                                                       f"\tThe report is empty: "
                                                       f"{str(report).replace('Unearned Revenue -', '').strip()}"}) \
                                            if gui_queue else None
                                        continue

                                await asyncio.sleep(5)
                                if report_data and "event" not in str(report).lower():
                                    read_data = Load_Report()
                                    output_status = read_data.read_report(report, report_data, self.locations,
                                                                          self.wb, self.output_file,
                                                                          self.report_date)
                                    if not output_status:
                                        self.error_log.append([report, None, "Error"])
                                        gui_queue.put({"status": f"\tError : "
                                                                 f"{str(report).replace('Unearned Revenue -', '').strip()}"}) \
                                            if gui_queue else None
                                        continue
                                    else:
                                        self.error_log.append([report, None, "Success"])
                                        gui_queue.put({"status": f"\tSuccess : "
                                                                 f"{str(report).replace('Unearned Revenue -', '').strip()}"}) \
                                            if gui_queue else None

                                elif "event" in str(report).lower() and event_dict:
                                    read_data = Load_Report()
                                    output_status = read_data.read_report(report, event_dict, self.locations,
                                                                          self.wb, self.output_file,
                                                                          self.report_date)
                                    if not output_status:
                                        self.error_log.append([report, None, "Error"])
                                        gui_queue.put({
                                            "status": f"\tError : "
                                                      f"{str(report).replace('Unearned Revenue -', '').strip()} - "
                                                      f"{location_name}"}) if gui_queue else None
                                        continue
                                    else:
                                        gui_queue.put(
                                            {"status": f"\tSuccess : "
                                                       f"{str(report).replace('Unearned Revenue -', '').strip()} - "
                                                       f"{location_name}"}) if gui_queue else None
                                else:
                                    self.error_log.append([report, None, "Error"])
                                    gui_queue.put({"status": f"\tError : "
                                                             f"{str(report).replace('Unearned Revenue -', '').strip()} -"
                                                             f" {location_name}"}) \
                                        if gui_queue else None
                                    continue
                            except Exception as e:
                                self.error_log.append([report, None, "Error"])
                                gui_queue.put({"status": f"\tError : "
                                                         f"{str(report).replace('Unearned Revenue -', '').strip()}"}) \
                                    if gui_queue else None
                                print(e)
                                continue

                    if self.GET_SEASON:
                        current_year = int(datetime.today().year)
                        season_years = [int(current_year), int(current_year - 1)]
                        season_years.append('future')
                        for year in season_years:
                            get_seasons = await self.get_season_list(year)
                        report = "Season"
                        read_data = Load_Report()
                        output_status = read_data.read_report(report, get_seasons, self.locations,
                                                              self.wb, self.output_file, self.report_date)
                        self.GET_SEASON = False

            else:
                self.LOGIN_FAILED = True
                print('\tUnable to Login!')
                gui_queue.put({'status': f'Error - Unable to Login!'}) if gui_queue else None
                return False

        return True

    def download_process(self, gui_queue, report, other_month):
        loop = asyncio.new_event_loop()
        executor = ThreadPoolExecutor(max_workers=3)
        future = asyncio.ensure_future(self.process_dash_report(gui_queue, report, other_month, executor), loop=loop)
        loop.run_until_complete(future)


class run_download_report:
    def __init__(self):
        self.gui_queue = None

    def run_download(self, report_date, output_path, client_name, all_reports, other_month):
        start_time = time.perf_counter()
        self.setting_xl = 'Dash_SettingSheet.xlsx'
        self.setting_wb = load_workbook(self.setting_xl)
        creds_values = self.setting_wb['Creds']
        credentials = [list(rows) for rows in creds_values.values]

        output_path = os.path.join(output_path, "Downloads", client_name)
        if not os.path.exists(output_path):
            os.makedirs(output_path)

        reportDate = str(report_date).replace('/', '-')
        output_file = os.path.join(output_path, f'Unearned Revenue - Summary {reportDate}.xlsx')

        error_file = f'error_file_{str(report_date).replace("/", "-")}.xlsx'
        if not os.path.isfile(error_file):
            error_wb = Workbook()
            error_sheet = error_wb.active
            error_sheet.append(["Report", "Location", "Status"])
        else:
            error_wb = load_workbook(error_file)
            error_sheet = error_wb.active

        dash = DASH_Download()
        dash.error_log = error_sheet
        dash.client_name = client_name
        dash.report_date = parse(report_date).strftime("%Y-%m-%d")
        dash.download_path = output_path
        dash.all_reports = all_reports
        dash.output_file = output_file
        dash.GET_SEASON = True

        if not os.path.isfile(output_file):
            dash.wb = Workbook()
            dash.ws = dash.wb.active
            dash.ws.title = "Summary"
            dash.wb.save(output_file)
        else:
            dash.wb = load_workbook(output_file)
            dash.ws = dash.wb.active

        dash.companyid = str([str(item[2]).strip() for item in credentials if
                              str(item[0]).strip().lower() == str(client_name).strip().lower()][0])
        dash.username = str([str(item[3]).strip() for item in credentials if
                             str(item[0]).strip().lower() == str(client_name).strip().lower()][0])
        dash.password = str([str(item[4]).strip() for item in credentials if
                             str(item[0]).strip().lower() == str(client_name).strip().lower()][0])

        if "Seasons" in dash.wb.sheetnames:
            sheet_index = dash.wb.get_sheet_by_name("Seasons")
            dash.wb.remove(sheet_index)
            dash.wb.save(output_file)
            time.sleep(3)

        for report_num, report in enumerate(all_reports):
            sheet_name = str(report).replace("Unearned Revenue -", "").strip()
            if sheet_name in dash.wb.sheetnames:
                sheet_index = dash.wb.get_sheet_by_name(sheet_name)
                dash.wb.remove(sheet_index)
                dash.wb.save(output_file)
                time.sleep(3)
            dash.wb.create_sheet(sheet_name)
            dash.download_process(self.gui_queue, report, other_month)
            if dash.LOGIN_FAILED:
                return False
            time.sleep(5)

        error_wb.save(error_file)
        end_time = time.perf_counter()
        time_taken = time.strftime("%H:%M:%S", time.gmtime(int(end_time - start_time)))
        self.gui_queue.put({'status': f'\n\nTime Taken : {time_taken}'}) if self.gui_queue else None
        return True
