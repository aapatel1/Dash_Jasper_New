from openpyxl import load_workbook
from time import sleep
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from dateutil.parser import parse
from datetime import datetime


class UnearnedSummary:
    def summary_report(self, report_date, output_path, client, summary_file):
        setting_xl = 'Dash_SettingSheet.xlsx'
        setting_wb = load_workbook(setting_xl)
        location_mapping = [list(row) for row in setting_wb['LocationMapping'].values]

        self.summaryTotal = 0
        summaryList = []
        summaryList.append([])
        summaryList.append([f"Unearned Revenue Summary as of {report_date}"])

        summary_wb = load_workbook(summary_file)
        summary_report_sheets = summary_wb.sheetnames
        if "SeasonNotInReport" in summary_report_sheets:
            remove_sht_no = summary_wb.get_sheet_by_name("SeasonNotInReport")
            summary_wb.remove(remove_sht_no)
            sleep(1)

        try:
            seasonList = summary_wb["Seasons"].values
            self.seasonList = [list(row) for row in seasonList]
        except:
            self.seasonList = []

        summary_report_sheets = summary_wb.sheetnames
        for report_sht in summary_report_sheets:
            if report_sht == "Seasons":
                continue

            if report_sht == "Summary":
                summary_x_data = [list(row) for row in summary_wb['Summary'].values]
                if summary_x_data:
                    summary_sheet_no = summary_wb.get_sheet_by_name("Summary")
                    summary_wb.remove(summary_sheet_no)
                    sleep(1)
                    summary_wb.create_sheet("Summary", 0)
                    sleep(1)
                    summary_wb.save(summary_file)
                continue
            else:
                report_data_sht = summary_wb[report_sht].values
                report_data = [list(row) for row in report_data_sht]
                if not report_data:
                    continue
                else:
                    summaryList.append([])
                    summaryList.append([f"{report_sht} - Unearned Revenue"])
                    summaryList.append(["GL Code", "Name", "Start date", "End Date", "Location", "Amount"])

            if not report_sht == "Membership":
                summary_dict = self.class_summary(report_data, report_sht, location_mapping)
                if summary_dict:
                    for key, value in summary_dict.items():
                        glCode = key[0]
                        seasonName = key[1]
                        start_date = key[2]
                        end_date = key[3]
                        location = key[4]
                        amount = value
                        summaryList.append([glCode, seasonName, start_date, end_date, location, amount])
                summaryList.append([])
            elif report_sht == "Membership":
                summary_dict = self.membership_summary(report_data, report_sht, location_mapping, report_date)
                if summary_dict:
                    for key, value in summary_dict.items():
                        summaryList.append([key[0], key[1], key[2], key[3], key[4], value])
                summaryList.append([])

        summaryList.append([])
        summaryList.append(["GRAND TOTAL", None, None, None, None, self.summaryTotal])

        summary_sheet = summary_wb['Summary']
        BORDER = False
        grand_total_row_num = None
        for row_num, row in enumerate(summaryList, 1):
            Yellow = False
            for col_num, value in enumerate(row, 2):
                if not Yellow:
                    try:
                        if parse(str(row[3])) <= parse(report_date):
                            Yellow = True
                    except:
                        pass
                summary_sheet.cell(row=row_num, column=col_num).value = value
                if row_num == 2:
                    summary_sheet.cell(row=row_num, column=col_num).alignment = \
                        Alignment(horizontal='center', vertical='center')
                    summary_sheet.cell(row=row_num, column=col_num).fill = PatternFill(start_color='FF00B0F0',
                                                                                       end_color='FF00B0F0',
                                                                                       fill_type="solid")
                    summary_sheet.cell(row=row_num, column=col_num).font = Font(name="Calibri", bold=True, size=18,
                                                                                color="FFFFFF")
                elif len(row) == 1 and "Unearned Revenue" in str(row[0]):
                    summary_sheet.merge_cells(start_row=row_num, end_row=row_num, start_column=2, end_column=7)
                    summary_sheet.cell(row=row_num, column=col_num).alignment = \
                        Alignment(horizontal='center', vertical='center')
                    summary_sheet.cell(row=row_num, column=col_num).font = Font(name="Calibri", bold=True, size=14,
                                                                                underline='single')
                elif "Total" in str(row[1]):
                    summary_sheet.cell(row=row_num, column=col_num).alignment = \
                        Alignment(horizontal='center', vertical='center')
                    summary_sheet.cell(row=row_num, column=col_num).font = Font(name="Calibri", bold=True, size=11)
                    summary_sheet.cell(row=row_num, column=col_num).border = \
                        Border(top=Side(border_style='thin', color='FF000000'),
                               right=Side(border_style='thin', color='FF000000'),
                               left=Side(border_style='thin', color='FF000000'),
                               bottom=Side(border_style='double', color='FF000000'))
                    BORDER = False
                elif "GL Code" in str(row[0]):
                    BORDER = True
                    summary_sheet.cell(row=row_num, column=col_num).alignment = \
                        Alignment(horizontal='center', vertical='center')
                    summary_sheet.cell(row=row_num, column=col_num).font = Font(name="Calibri", bold=True, size=11)
                    summary_sheet.cell(row=row_num, column=col_num).fill = PatternFill(start_color='FFFCE4D6',
                                                                                       end_color='FFFCE4D6',
                                                                                       fill_type="solid")
                elif "GRAND TOTAL" in str(row[0]):
                    grand_total_row_num = row_num
                    summary_sheet.cell(row=row_num, column=col_num).alignment = \
                        Alignment(horizontal='center', vertical='center')
                    summary_sheet.cell(row=row_num, column=col_num).font = Font(name="Calibri", bold=True, size=11)
                    summary_sheet.cell(row=row_num, column=col_num).border = \
                        Border(top=Side(border_style='thin', color='FF000000'),
                               right=Side(border_style='thin', color='FF000000'),
                               left=Side(border_style='thin', color='FF000000'),
                               bottom=Side(border_style='double', color='FF000000'))
                if Yellow:
                    summary_sheet.cell(row=row_num, column=col_num).fill = PatternFill(start_color='FFFBFF42',
                                                                                       end_color='FFFBFF42',
                                                                                       fill_type="solid")
                if col_num == 7:
                    summary_sheet.cell(row=row_num, column=col_num).number_format = f'$#,##0.00_);[Red]($#,##0.00)'
                if BORDER:
                    summary_sheet.cell(row=row_num, column=col_num).border = \
                        Border(top=Side(border_style='thin', color='FF000000'),
                               right=Side(border_style='thin', color='FF000000'),
                               left=Side(border_style='thin', color='FF000000'),
                               bottom=Side(border_style='thin', color='FF000000'))

                try:
                    if col_num == 4 or col_num == 5:
                        dateee = parse(str(value))
                        summary_sheet.cell(row=row_num, column=col_num).value = dateee
                        summary_sheet.cell(row=row_num, column=col_num).number_format = f'mm-dd-yyyy'
                except:
                    pass

        summary_sheet.merge_cells(start_row=2, end_row=2, start_column=2, end_column=7)
        summary_sheet.merge_cells(start_row=grand_total_row_num, end_row=grand_total_row_num,
                                  start_column=2, end_column=6)

        summary_sheet.column_dimensions[get_column_letter(2)].width = 15
        summary_sheet.column_dimensions[get_column_letter(3)].width = 50
        summary_sheet.column_dimensions[get_column_letter(4)].width = 15
        summary_sheet.column_dimensions[get_column_letter(5)].width = 15
        summary_sheet.column_dimensions[get_column_letter(6)].width = 20
        summary_sheet.column_dimensions[get_column_letter(7)].width = 15

        summary_wb.save(summary_file)

    def class_summary(self, report_data, report_sht, location_mapping):
        location_col = start_date_col = end_date_col = season_col = gl_col = unearned_col = None
        for col_num, value in enumerate(report_data[0]):
            if value == "Location":
                location_col = col_num
            elif value == "Start Date":
                start_date_col = col_num
            elif value == "End Date":
                end_date_col = col_num
            elif value == "Season Name" or value == "GLCode Description":
                season_col = col_num
            elif value == "GL Code":
                gl_col = col_num
            elif value == "Unearned Revenue Closing" or value == "Amount":
                unearned_col = col_num

        summary_dict = {}
        class_total = 0
        for row in report_data[1:]:
            gl_code = row[gl_col]
            season_name = row[season_col]
            unearned_amount = row[unearned_col]
            location = start_date = end_date = None
            if report_sht == "Class" or report_sht == "Team" or report_sht == "Camp":
                location = row[location_col]
                start_date = row[start_date_col]
                end_date = row[end_date_col]
            elif report_sht == "Event & Rental":
                location = row[location_col]

            if not unearned_amount:
                continue
            class_total += round(unearned_amount, 2)
            if not summary_dict:
                summary_dict[gl_code, season_name, start_date, end_date, location] = round(unearned_amount, 2)
            else:
                if (gl_code, season_name, start_date, end_date, location) not in summary_dict.keys():
                    summary_dict[gl_code, season_name, start_date, end_date, location] = round(unearned_amount, 2)
                else:
                    summary_dict[gl_code, season_name, start_date, end_date, location] += round(unearned_amount, 2)

        self.summaryTotal += class_total
        summary_dict[None, "Total", None, None, None] = round(class_total, 2)
        return summary_dict

    def membership_summary(self, report_data, report_sht, location_mapping, report_date):
        start_line = report_month_col = None
        summary_dict = {}
        class_total = 0
        for line_num, line in enumerate(report_data):
            if "Unearned Membership" in line:
                start_line = line_num
                for col_num, value in enumerate(report_data[start_line+1]):
                    if col_num >= 2:
                        if parse(report_date).strftime("%b-%Y") == value:
                            report_month_col = col_num
                            break
            if report_month_col:
                for row in report_data[start_line+2:]:
                    if row[report_month_col]:
                        class_total += row[report_month_col]
                        summary_dict[row[1], row[0], None, None, None] = row[report_month_col]

            if summary_dict:
                self.summaryTotal += class_total
                summary_dict[None, "Total", None, None, None] = round(class_total, 2)
                return summary_dict

    def season_matching(self, report_date, output_path, client, summary_file):
        summary_wb = load_workbook(summary_file)
        summary_report_sheets = summary_wb.sheetnames

        season_sheet = summary_wb["Seasons"].values
        season_data = [list(row) for row_num, row in enumerate(season_sheet) if row_num > 0
                       if parse(str(row[2])) > parse(str(report_date))]
        if not season_data:
            return False

        err_summary = []
        for sum_sht in summary_report_sheets:
            if str(sum_sht) == "Camp":
                camp_sheet = summary_wb[sum_sht].values
                camp_data = [list(row) for row in camp_sheet]
                if not camp_data:
                    continue
                camp_seasons = [item[3] for item in camp_data[1:]]
                for season_item in season_data:
                    if "Per-session" in season_item:
                        season_name = " ".join(str(season_item[0]).split()[:-1]).strip()
                        if str(season_name) not in camp_seasons:
                            err_summary.append(season_item)

            if str(sum_sht) == "Class":
                class_sheet = summary_wb[sum_sht].values
                class_data = [list(row) for row in class_sheet]
                if not class_data:
                    continue
                camp_seasons = [item[3] for item in class_data[1:]]
                for season_item in season_data:
                    if "Class" in season_item:
                        season_name = " ".join(str(season_item[0]).split()[:-1]).strip()
                        if str(season_name) not in camp_seasons:
                            err_summary.append(season_item)

            if str(sum_sht) == "Team":
                team_sheet = summary_wb[sum_sht].values
                team_data = [list(row) for row in team_sheet]
                if not team_data:
                    continue
                camp_seasons = [item[3] for item in team_data[1:]]
                for season_item in season_data:
                    if "League" in season_item:
                        season_name = " ".join(str(season_item[0]).split()[:-1]).strip()
                        if str(season_name) not in camp_seasons:
                            err_summary.append(season_item)

        if err_summary:
            err_summary_sht = summary_wb.create_sheet("SeasonNotInReport")
            err_summary_sht.append(["Season", "Start Date", "End Date", "Type", "Links"])
            for row in err_summary:
                err_summary_sht.append(row)

        summary_wb.save(summary_file)


class run_combined_summary:
    def __init__(self):
        self.gui_queue = None

    def run(self, report_date, output_path, client, summary_file):
        try:
            summ = UnearnedSummary()
            summ.summary_report(report_date, output_path, client, summary_file)
            summ.season_matching(report_date, output_path, client, summary_file)
            self.gui_queue.put({'status': f'\nSummary Processed.'}) if self.gui_queue else None
        except Exception as e:
            self.gui_queue.put({'status': f'\nFailed : {str(e)}'}) if self.gui_queue else None












# for line in location_mapping:
#     if gl_code in line:
#         location = line[3]
#         break
# if not location:
#     try:
#         season_loc = str(row[season_col]).split("(")[1].replace(")", "").strip() or None
#         for line in location_mapping:
#             if line[2] == season_loc:
#                 location = line[3]
#                 break
#     except:
#         location = None